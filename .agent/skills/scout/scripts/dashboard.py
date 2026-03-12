#!/usr/bin/env python3
from __future__ import annotations
"""
dashboard.py — Reads data/influencers.xlsx → writes data/dashboard.html.

Self-contained HTML dashboard with influencer cards, candidates table,
similar users table, and outreach log.
"""
import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import excel

OUTPUT_PATH = excel.PROJECT_ROOT / 'data' / 'dashboard.html'


def _read_sheet(sheet_name: str) -> list[dict]:
    if not excel.XLSX_PATH.exists():
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel.XLSX_PATH, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet_name]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {}
            for i, h in enumerate(headers, 1):
                if h:
                    v = ws.cell(row=r, column=i).value
                    row[h] = str(v) if v is not None else ''
            if any(row.values()):
                rows.append(row)
        wb.close()
        return rows
    except Exception as e:
        print(f"[dashboard] warning reading {sheet_name}: {e}")
        return []


def generate(campaign: str | None = None):
    influencers = _read_sheet('Influencers')
    candidates = _read_sheet('Candidates')
    similar = _read_sheet('Similar Users')
    outreach = _read_sheet('Outreach')
    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M')

    # Filter by campaign if specified
    if campaign:
        influencers = [r for r in influencers if r.get('campaign') == campaign]
        candidates = [r for r in candidates if r.get('campaign') == campaign]
        outreach = [r for r in outreach if r.get('campaign') == campaign]

    data_json = json.dumps({
        'influencers': influencers,
        'candidates': candidates,
        'similar': similar,
        'outreach': outreach,
        'generated_at': generated_at,
    }, ensure_ascii=False, indent=2)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Influencer Scout Dashboard</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

  :root {{
    --bg: #f2f2f7;
    --surface: #ffffff;
    --surface2: #f9f9fb;
    --border: #e4e4eb;
    --accent: #0071e3;
    --accent-light: #e8f1fb;
    --text: #1c1c1e;
    --text2: #6e6e73;
    --text3: #aeaeb2;
    --green: #34c759;
    --red: #ff3b30;
    --yellow: #ff9f0a;
    --radius: 14px;
    --radius-sm: 8px;
    --shadow: 0 1px 3px rgba(0,0,0,.07), 0 4px 12px rgba(0,0,0,.04);
  }}

  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'SF Pro Text', 'Segoe UI', sans-serif;
    background: var(--bg);
    color: var(--text);
    min-height: 100vh;
    font-size: 14px;
  }}

  .topbar {{
    position: sticky; top: 0; z-index: 200;
    background: rgba(255,255,255,.85);
    backdrop-filter: saturate(180%) blur(20px);
    -webkit-backdrop-filter: saturate(180%) blur(20px);
    border-bottom: 1px solid var(--border);
    padding: 10px 24px;
    display: flex; align-items: center; gap: 12px; flex-wrap: wrap;
  }}
  .topbar-brand {{
    display: flex; align-items: center; gap: 8px; flex: 1; min-width: 140px;
  }}
  .topbar-brand .logo {{
    width: 28px; height: 28px; background: var(--accent);
    border-radius: 8px; display: flex; align-items: center; justify-content: center;
    color: #fff; font-size: 14px; font-weight: 800; letter-spacing: -.5px;
  }}
  .topbar-brand h1 {{ font-size: 16px; font-weight: 700; color: var(--text); }}
  .topbar-brand .gen-time {{ font-size: 11px; color: var(--text3); margin-left: 4px; }}

  .filters {{ display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }}
  .filter-select {{
    padding: 6px 10px; border: 1px solid var(--border);
    border-radius: var(--radius-sm); font-size: 13px;
    background: var(--surface2); color: var(--text);
    cursor: pointer; appearance: none;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath fill='%236e6e73' d='M6 8L1 3h10z'/%3E%3C/svg%3E");
    background-repeat: no-repeat; background-position: right 8px center;
    padding-right: 26px;
  }}
  .filter-select:focus {{ outline: none; border-color: var(--accent); }}

  .tabs {{ display: flex; gap: 4px; background: var(--bg); border-radius: 10px; padding: 3px; }}
  .tab {{
    padding: 5px 14px; border-radius: 8px; border: none;
    background: transparent; font-size: 13px; font-weight: 500;
    color: var(--text2); cursor: pointer; transition: all .15s;
  }}
  .tab.active {{
    background: var(--surface); color: var(--text);
    box-shadow: 0 1px 3px rgba(0,0,0,.12);
  }}

  .content {{ padding: 20px 24px; max-width: 1280px; margin: 0 auto; }}
  .section {{ display: none; }}
  .section.active {{ display: block; }}

  .summary-bar {{ display: flex; gap: 12px; margin-bottom: 20px; flex-wrap: wrap; }}
  .stat-chip {{
    background: var(--surface); border-radius: var(--radius-sm);
    padding: 10px 16px; box-shadow: var(--shadow);
    display: flex; flex-direction: column; gap: 2px;
  }}
  .stat-chip .label {{ font-size: 11px; color: var(--text3); text-transform: uppercase; letter-spacing: .5px; }}
  .stat-chip .value {{ font-size: 20px; font-weight: 700; color: var(--text); }}

  .controls {{ display: flex; align-items: center; gap: 10px; margin-bottom: 16px; flex-wrap: wrap; }}
  .controls label {{ font-size: 13px; color: var(--text2); }}

  .grid {{ display: grid; gap: 14px; grid-template-columns: repeat(auto-fill, minmax(240px, 1fr)); }}

  .card {{
    background: var(--surface); border-radius: var(--radius);
    padding: 16px; box-shadow: var(--shadow);
    display: flex; flex-direction: column; gap: 0;
    transition: transform .15s, box-shadow .15s;
  }}
  .card:hover {{ transform: translateY(-2px); box-shadow: 0 4px 16px rgba(0,0,0,.10); }}
  .card-header {{ display: flex; align-items: flex-start; justify-content: space-between; margin-bottom: 12px; }}
  .card .handle {{ font-weight: 700; font-size: 15px; line-height: 1.2; color: var(--text); }}
  .card .handle a {{ color: inherit; text-decoration: none; }}
  .card .handle a:hover {{ color: var(--accent); }}

  .card .views-grid {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 6px; margin-bottom: 12px; }}
  .card .view-stat {{ background: var(--surface2); border-radius: 8px; padding: 6px 8px; text-align: center; }}
  .card .view-stat .v-label {{ font-size: 10px; color: var(--text3); text-transform: uppercase; letter-spacing: .3px; }}
  .card .view-stat .v-value {{ font-size: 14px; font-weight: 700; color: var(--accent); }}

  .card .tags {{ display: flex; flex-wrap: wrap; gap: 4px; margin-bottom: 10px; }}
  .tag {{ display: inline-block; border-radius: 6px; padding: 2px 7px; font-size: 11px; font-weight: 500; }}
  .tag-kw {{ background: var(--accent-light); color: var(--accent); }}
  .tag-campaign {{ background: #f0f0f5; color: #555; }}

  .card .meta {{ font-size: 11px; color: var(--text3); margin-bottom: 6px; }}
  .card .meta a {{ color: var(--accent); text-decoration: none; }}
  .card .meta a:hover {{ text-decoration: underline; }}
  .card .date {{ font-size: 11px; color: var(--text3); margin-bottom: 10px; }}
  .card .cta {{
    display: block; text-align: center; padding: 7px;
    background: var(--accent); color: #fff;
    border-radius: var(--radius-sm); font-size: 12px; font-weight: 600;
    text-decoration: none; margin-top: auto; transition: background .15s;
  }}
  .card .cta:hover {{ background: #005bbf; }}

  .empty {{ text-align: center; padding: 60px 20px; color: var(--text3); font-size: 14px; }}

  .table-wrap {{ overflow-x: auto; border-radius: var(--radius); box-shadow: var(--shadow); }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; background: var(--surface); }}
  thead tr {{ background: var(--surface2); }}
  th {{
    text-align: left; padding: 11px 14px; font-weight: 600; color: var(--text2);
    font-size: 12px; text-transform: uppercase; letter-spacing: .4px;
    border-bottom: 1px solid var(--border); white-space: nowrap;
  }}
  td {{ padding: 9px 14px; border-bottom: 1px solid var(--border); }}
  tbody tr:last-child td {{ border-bottom: none; }}
  tbody tr:hover {{ background: var(--surface2); }}
  td a {{ color: var(--accent); text-decoration: none; }}
  td a:hover {{ text-decoration: underline; }}

  .badge {{ display: inline-block; border-radius: 20px; padding: 2px 9px; font-size: 11px; font-weight: 600; }}
  .badge-pending {{ background: #f0f0f5; color: #888; }}
  .badge-qualified {{ background: #d4f5de; color: #1a7a36; }}
  .badge-not_qualified {{ background: #fde8e8; color: #c0392b; }}
  .badge-error {{ background: #fff3cd; color: #856404; }}
  .badge-sent {{ background: #d4f5de; color: #1a7a36; }}
  .badge-failed {{ background: #fde8e8; color: #c0392b; }}
  .play-count {{ font-variant-numeric: tabular-nums; }}

  footer {{ text-align: center; font-size: 11px; color: var(--text3); padding: 24px; margin-top: 8px; }}
</style>
</head>
<body>

<div class="topbar">
  <div class="topbar-brand">
    <div class="logo">IS</div>
    <h1>Scout Dashboard</h1>
    <span class="gen-time">Updated {generated_at}</span>
  </div>
  <div class="filters">
    <select class="filter-select" id="campaignFilter" onchange="applyFilters()">
      <option value="">All Campaigns</option>
    </select>
    <select class="filter-select" id="keywordFilter" onchange="applyFilters()">
      <option value="">All Keywords</option>
    </select>
  </div>
  <div class="tabs">
    <button class="tab active" onclick="showTab('influencers', this)">Influencers</button>
    <button class="tab" onclick="showTab('candidates', this)">Candidates</button>
    <button class="tab" onclick="showTab('similar', this)">Similar</button>
    <button class="tab" onclick="showTab('outreach', this)">Outreach</button>
  </div>
</div>

<div class="content">

  <div id="influencers" class="section active">
    <div class="summary-bar" id="influencer-summary"></div>
    <div class="controls">
      <label>Sort by</label>
      <select class="filter-select" id="influencer-sort" onchange="renderInfluencers()">
        <option value="max_views">Max Views</option>
        <option value="median_views">Median Views</option>
        <option value="scouted_date">Date Scouted</option>
      </select>
    </div>
    <div class="grid" id="influencer-grid"></div>
  </div>

  <div id="candidates" class="section">
    <div class="summary-bar" id="candidate-summary"></div>
    <div class="table-wrap">
      <table>
        <thead><tr><th>Handle</th><th>Keyword</th><th>Campaign</th><th>Status</th><th>Play Count</th></tr></thead>
        <tbody id="candidate-tbody"></tbody>
      </table>
    </div>
  </div>

  <div id="similar" class="section">
    <div class="table-wrap">
      <table>
        <thead><tr><th>Queried</th><th>Similar</th><th>Bio Link</th><th>Emails</th><th>Date</th></tr></thead>
        <tbody id="similar-tbody"></tbody>
      </table>
    </div>
  </div>

  <div id="outreach" class="section">
    <div class="table-wrap">
      <table>
        <thead><tr><th>Handle</th><th>Emails</th><th>Subject</th><th>Campaign</th><th>Sent At</th><th>Status</th></tr></thead>
        <tbody id="outreach-tbody"></tbody>
      </table>
    </div>
  </div>

</div>

<footer>Influencer Scout &middot; {generated_at}</footer>

<script>
const DATA = {data_json};

function fmt(n) {{
  n = parseInt(n) || 0;
  if (n >= 1e6) return (n/1e6).toFixed(1) + 'M';
  if (n >= 1e3) return (n/1e3).toFixed(1) + 'K';
  return n.toString();
}}

function esc(s) {{ return (s||'').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }}

const allCampaigns = [...new Set([
  ...DATA.influencers.map(r => r.campaign),
  ...DATA.candidates.map(r => r.campaign),
].filter(Boolean))].sort();

const campaignSel = document.getElementById('campaignFilter');
allCampaigns.forEach(c => {{
  const o = document.createElement('option');
  o.value = c; o.textContent = c;
  campaignSel.appendChild(o);
}});

function getCampaign() {{ return document.getElementById('campaignFilter').value; }}
function getKeyword()  {{ return document.getElementById('keywordFilter').value; }}
function getSortKey()  {{ return document.getElementById('influencer-sort').value; }}

function updateKeywordFilter(rows) {{
  const current = getKeyword();
  const keywords = [...new Set(rows.map(r => r.keyword).filter(Boolean))].sort();
  const sel = document.getElementById('keywordFilter');
  sel.innerHTML = '<option value="">All Keywords</option>';
  keywords.forEach(k => {{
    const o = document.createElement('option');
    o.value = k; o.textContent = '#' + k;
    if (k === current) o.selected = true;
    sel.appendChild(o);
  }});
}}

function applyFilters() {{ renderInfluencers(); renderCandidates(); renderSimilar(); renderOutreach(); }}

function renderInfluencers() {{
  const campaign = getCampaign(), keyword = getKeyword(), sortKey = getSortKey();
  let rows = DATA.influencers.filter(r =>
    (!campaign || r.campaign === campaign) && (!keyword || r.keyword === keyword)
  );
  rows.sort((a, b) => {{
    if (sortKey === 'scouted_date') return (b.scouted_date || '').localeCompare(a.scouted_date || '');
    return (parseInt(b[sortKey]) || 0) - (parseInt(a[sortKey]) || 0);
  }});
  updateKeywordFilter(DATA.influencers.filter(r => !campaign || r.campaign === campaign));

  const withEmails = rows.filter(r => r.emails).length;
  document.getElementById('influencer-summary').innerHTML = `
    <div class="stat-chip"><span class="label">Total</span><span class="value">${{rows.length}}</span></div>
    <div class="stat-chip"><span class="label">With Email</span><span class="value">${{withEmails}}</span></div>
  `;

  const grid = document.getElementById('influencer-grid');
  if (!rows.length) {{
    grid.innerHTML = '<div class="empty" style="grid-column:1/-1">No influencers match the current filters.</div>';
    return;
  }}
  grid.innerHTML = rows.map(r => {{
    const emailLinks = (r.emails || '').split(',').filter(Boolean).map(e =>
      `<a href="mailto:${{e.trim()}}">${{esc(e.trim())}}</a>`
    ).join(', ');
    const bioLink = r.bio_link ? `<a href="${{esc(r.bio_link)}}" target="_blank">${{esc(r.bio_link)}}</a>` : '';
    return `
    <div class="card">
      <div class="card-header">
        <div class="handle"><a href="${{r.profile_url || 'https://www.tiktok.com/@'+r.handle}}" target="_blank">@${{esc(r.handle)}}</a></div>
      </div>
      <div class="views-grid">
        <div class="view-stat"><div class="v-label">Max</div><div class="v-value">${{fmt(r.max_views)}}</div></div>
        <div class="view-stat"><div class="v-label">Med</div><div class="v-value">${{fmt(r.median_views)}}</div></div>
        <div class="view-stat"><div class="v-label">Min</div><div class="v-value">${{fmt(r.min_views)}}</div></div>
      </div>
      <div class="tags">
        ${{r.keyword  ? `<span class="tag tag-kw">#${{esc(r.keyword)}}</span>` : ''}}
        ${{r.campaign ? `<span class="tag tag-campaign">${{esc(r.campaign)}}</span>` : ''}}
      </div>
      ${{emailLinks ? `<div class="meta">${{emailLinks}}</div>` : ''}}
      ${{bioLink ? `<div class="meta">${{bioLink}}</div>` : ''}}
      <div class="date">${{r.scouted_date || ''}}</div>
      <a class="cta" href="${{r.profile_url || 'https://www.tiktok.com/@'+r.handle}}" target="_blank">View on TikTok</a>
    </div>`;
  }}).join('');
}}

function renderCandidates() {{
  const campaign = getCampaign(), keyword = getKeyword();
  let rows = DATA.candidates.filter(r =>
    (!campaign || r.campaign === campaign) && (!keyword || r.keyword === keyword)
  );
  const qualified = rows.filter(r => r.audit_status === 'qualified').length;
  const notQual = rows.filter(r => r.audit_status === 'not_qualified').length;
  const pending = rows.filter(r => !r.audit_status || r.audit_status === 'pending').length;
  document.getElementById('candidate-summary').innerHTML = `
    <div class="stat-chip"><span class="label">Total</span><span class="value">${{rows.length}}</span></div>
    <div class="stat-chip"><span class="label">Qualified</span><span class="value">${{qualified}}</span></div>
    <div class="stat-chip"><span class="label">Not Qualified</span><span class="value">${{notQual}}</span></div>
    <div class="stat-chip"><span class="label">Pending</span><span class="value">${{pending}}</span></div>
  `;
  const tbody = document.getElementById('candidate-tbody');
  if (!rows.length) {{ tbody.innerHTML = '<tr><td colspan="5" class="empty">No candidates.</td></tr>'; return; }}
  tbody.innerHTML = rows.map(r => {{
    const status = r.audit_status || 'pending';
    return `<tr>
      <td><a href="https://www.tiktok.com/@${{esc(r.handle)}}" target="_blank">@${{esc(r.handle)}}</a></td>
      <td>${{r.keyword ? `<span class="tag tag-kw">#${{esc(r.keyword)}}</span>` : ''}}</td>
      <td>${{esc(r.campaign)}}</td>
      <td><span class="badge badge-${{status}}">${{status.replace('_',' ')}}</span></td>
      <td class="play-count">${{fmt(r.triggering_play_count)}}</td>
    </tr>`;
  }}).join('');
}}

function renderSimilar() {{
  const rows = DATA.similar;
  const tbody = document.getElementById('similar-tbody');
  if (!rows.length) {{ tbody.innerHTML = '<tr><td colspan="5" class="empty">No similar user lookups yet.</td></tr>'; return; }}
  tbody.innerHTML = rows.map(r => `<tr>
    <td><a href="https://www.tiktok.com/@${{esc(r.queried_handle)}}" target="_blank">@${{esc(r.queried_handle)}}</a></td>
    <td><a href="${{esc(r.profile_url)}}" target="_blank">@${{esc(r.similar_handle)}}</a></td>
    <td>${{r.bio_link ? `<a href="${{esc(r.bio_link)}}" target="_blank">${{esc(r.bio_link)}}</a>` : ''}}</td>
    <td>${{esc(r.emails)}}</td>
    <td>${{esc(r.lookup_date)}}</td>
  </tr>`).join('');
}}

function renderOutreach() {{
  const campaign = getCampaign();
  let rows = DATA.outreach.filter(r => !campaign || r.campaign === campaign);
  const tbody = document.getElementById('outreach-tbody');
  if (!rows.length) {{ tbody.innerHTML = '<tr><td colspan="6" class="empty">No outreach sent yet.</td></tr>'; return; }}
  tbody.innerHTML = rows.map(r => `<tr>
    <td>@${{esc(r.handle)}}</td>
    <td>${{esc(r.emails)}}</td>
    <td>${{esc(r.subject)}}</td>
    <td>${{esc(r.campaign)}}</td>
    <td>${{esc(r.sent_at)}}</td>
    <td><span class="badge badge-${{r.status || 'pending'}}">${{esc(r.status)}}</span></td>
  </tr>`).join('');
}}

function showTab(name, btn) {{
  document.querySelectorAll('.section').forEach(s => s.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.getElementById(name).classList.add('active');
  btn.classList.add('active');
}}

renderInfluencers();
renderCandidates();
renderSimilar();
renderOutreach();
</script>
</body>
</html>"""

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(html, encoding='utf-8')
    print(f"[dashboard] Generated: {OUTPUT_PATH}")
    return OUTPUT_PATH


if __name__ == '__main__':
    generate()
