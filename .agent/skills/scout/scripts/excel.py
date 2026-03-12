#!/usr/bin/env python3
"""
excel.py — xlsx read/write helpers for influencer scouting.
All functions operate on data/influencers.xlsx relative to project root.
"""
from __future__ import annotations

import logging
import re
import subprocess
import time
from datetime import date
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

log = logging.getLogger('scout.excel')

# ---------------------------------------------------------------------------
# Sheet schemas
# ---------------------------------------------------------------------------
CANDIDATES_COLS = [
    'handle', 'triggering_video_url', 'triggering_play_count',
    'keyword', 'campaign', 'audit_status',
]
INFLUENCERS_COLS = [
    'handle', 'profile_url', 'max_views', 'min_views', 'median_views',
    'triggering_video_url', 'triggering_play_count',
    'keyword', 'campaign', 'scouted_date',
    'bio', 'bio_link', 'emails', 'notes',
]
SEARCH_LOG_COLS = [
    'keyword', 'results_checked', 'candidates_found', 'qualified',
    'duration_mins', 'campaign', 'run_date',
]
SIMILAR_USERS_COLS = [
    'queried_handle', 'similar_handle', 'profile_url',
    'bio', 'bio_link', 'emails', 'lookup_date', 'requested_by',
]
OUTREACH_COLS = [
    'handle', 'emails', 'subject', 'campaign', 'sent_at', 'status',
]

PROJECT_ROOT = Path(__file__).resolve().parents[4]
XLSX_PATH = PROJECT_ROOT / 'data' / 'influencers.xlsx'


def _close_xlsx_in_app() -> bool:
    """Try to gracefully close the xlsx if open in Excel or Numbers on macOS.
    Returns True if successfully closed (or wasn't actually locked),
    False if the file is still locked after our attempt."""
    lock_file = XLSX_PATH.parent / f'~${XLSX_PATH.name}'
    if not lock_file.exists():
        return True

    # Find which app holds the file open
    app_name = None
    try:
        result = subprocess.run(
            ['lsof', str(XLSX_PATH)],
            capture_output=True, text=True, timeout=5,
        )
        for line in result.stdout.splitlines()[1:]:  # skip header
            parts = line.split()
            if parts:
                cmd = parts[0]
                if 'Excel' in cmd or 'excel' in cmd:
                    app_name = 'Microsoft Excel'
                    break
                if 'Numbers' in cmd or 'numbers' in cmd:
                    app_name = 'Numbers'
                    break
    except Exception:
        pass

    if not app_name:
        # Lock file exists but can't identify the app — try both
        for name in ('Microsoft Excel', 'Numbers'):
            _applescript_close(name)
    else:
        _applescript_close(app_name)

    # Wait briefly for the app to release the file
    for _ in range(6):
        time.sleep(0.5)
        if not lock_file.exists():
            log.info("Closed %s in %s before writing.", XLSX_PATH.name, app_name or 'spreadsheet app')
            return True

    return False


def _applescript_close(app_name: str):
    """Ask a macOS app to save and close our workbook via AppleScript."""
    filename = XLSX_PATH.name
    if app_name == 'Microsoft Excel':
        script = f'''
            tell application "Microsoft Excel"
                repeat with wb in workbooks
                    if name of wb is "{filename}" then
                        close wb saving yes
                        exit repeat
                    end if
                end repeat
            end tell
        '''
    else:
        # Numbers
        script = f'''
            tell application "Numbers"
                repeat with doc in documents
                    if name of doc contains "{XLSX_PATH.stem}" then
                        close doc saving yes
                        exit repeat
                    end if
                end repeat
            end tell
        '''
    try:
        subprocess.run(
            ['osascript', '-e', script],
            capture_output=True, timeout=10,
        )
    except Exception as e:
        log.debug("AppleScript close failed for %s: %s", app_name, e)


def _ensure_xlsx_writable():
    """Ensure the xlsx is not open in another app. Auto-closes if possible."""
    lock_file = XLSX_PATH.parent / f'~${XLSX_PATH.name}'
    if not lock_file.exists():
        return
    log.warning("%s appears open in a spreadsheet app. Attempting to close...", XLSX_PATH.name)
    if not _close_xlsx_in_app():
        raise RuntimeError(
            f"{XLSX_PATH.name} is open in another application and could not be "
            f"closed automatically. Please close it and try again."
        )


def _style_header(cell):
    cell.font = Font(bold=True)
    cell.fill = PatternFill('solid', start_color='D9E1F2')
    cell.alignment = Alignment(horizontal='center')


def _init_xlsx():
    """Create influencers.xlsx with all sheets from scratch."""
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    def make(name, cols, widths):
        ws = wb.create_sheet(name)
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=h)
            _style_header(c)
        for i, w in enumerate(widths):
            ws.column_dimensions[ws.cell(row=1, column=i + 1).column_letter].width = w
        return ws

    make('Influencers', INFLUENCERS_COLS,
         [22, 45, 14, 14, 14, 50, 22, 30, 20, 14, 30, 30, 30, 30])
    make('Candidates', CANDIDATES_COLS,
         [22, 50, 16, 30, 20, 16])
    make('Search Log', SEARCH_LOG_COLS,
         [30, 18, 18, 12, 16, 20, 14])
    make('Similar Users', SIMILAR_USERS_COLS,
         [20, 20, 40, 30, 30, 30, 14, 20])
    make('Outreach', OUTREACH_COLS,
         [20, 30, 40, 20, 20, 12])

    wb.save(XLSX_PATH)
    wb.close()
    print(f'Created {XLSX_PATH}')


def _get_wb_ws(sheet_name: str, required_cols: list):
    """Open xlsx and return (wb, ws). Creates xlsx/sheet with headers if missing."""
    _ensure_xlsx_writable()
    if not XLSX_PATH.exists():
        _init_xlsx()
    wb = load_workbook(XLSX_PATH)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        for i, col in enumerate(required_cols, 1):
            c = ws.cell(row=1, column=i, value=col)
            _style_header(c)
    else:
        ws = wb[sheet_name]
        _ensure_columns(ws, required_cols)

    return wb, ws


def _ensure_columns(ws, required_cols: list):
    """Append missing columns to header row — never drops existing."""
    if ws.max_row == 0:
        return
    existing = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for col_name in required_cols:
        if col_name not in existing:
            next_col = ws.max_column + 1
            c = ws.cell(row=1, column=next_col, value=col_name)
            _style_header(c)
            existing.append(col_name)


def _col_index(ws, name: str):
    """Return 1-based column index by header name, or None."""
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == name:
            return c
    return None


def _row_to_dict(ws, row_num: int) -> dict:
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    return {h: ws.cell(row=row_num, column=i + 1).value for i, h in enumerate(headers) if h}


def _append_row(ws, headers, row: dict):
    next_row = ws.max_row + 1
    for i, h in enumerate(headers, 1):
        if h and h in row:
            ws.cell(row=next_row, column=i, value=row[h])


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

DEFAULTS = {
    'view_threshold': 10000,
    'min_video_views': 10000,
    'recent_video_count': 10,
    'max_candidates_per_keyword': 5,
}


def load_config(campaign_name: str) -> dict:
    """Parse campaign.md YAML front-matter, apply hardcoded defaults."""
    config = {}

    campaign_path = PROJECT_ROOT / 'context' / 'campaigns' / campaign_name / 'campaign.md'
    if campaign_path.exists():
        text = campaign_path.read_text()
        if text.startswith('---'):
            fm_block = text.split('---')[1]
            for line in fm_block.splitlines():
                line = line.strip()
                if ':' in line and not line.startswith('#'):
                    key, _, val = line.partition(':')
                    val = val.strip().strip('"').strip("'")
                    if val.lstrip('-').isdigit():
                        config[key.strip()] = int(val)
                    elif val.replace('.', '', 1).isdigit():
                        config[key.strip()] = float(val)
                    elif val:
                        config[key.strip()] = val

    for k, v in DEFAULTS.items():
        config.setdefault(k, v)
    return config


# ---------------------------------------------------------------------------
# Candidates
# ---------------------------------------------------------------------------

def append_candidates(rows: list[dict]) -> int:
    """Append rows to Candidates sheet; skip if handle+campaign already exists."""
    wb, ws = _get_wb_ws('Candidates', CANDIDATES_COLS)
    existing = set()
    handle_col = _col_index(ws, 'handle') or 1
    camp_col = _col_index(ws, 'campaign') or 5
    for r in range(2, ws.max_row + 1):
        h = ws.cell(row=r, column=handle_col).value
        camp = ws.cell(row=r, column=camp_col).value
        if h and camp:
            existing.add((h, camp))

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    added = 0
    for row in rows:
        key = (row.get('handle'), row.get('campaign'))
        if key in existing:
            continue
        _append_row(ws, headers, row)
        existing.add(key)
        added += 1

    wb.save(XLSX_PATH)
    wb.close()
    return added


def candidate_exists(handle: str, campaign: str) -> bool:
    if not XLSX_PATH.exists():
        return False
    wb = load_workbook(XLSX_PATH, read_only=True)
    if 'Candidates' not in wb.sheetnames:
        wb.close()
        return False
    ws = wb['Candidates']
    handle_col = _col_index(ws, 'handle') or 1
    camp_col = _col_index(ws, 'campaign') or 5
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=handle_col).value == handle and \
           ws.cell(row=r, column=camp_col).value == campaign:
            wb.close()
            return True
    wb.close()
    return False


def update_candidate_status(handle: str, campaign: str, status: str, notes: str = ''):
    wb, ws = _get_wb_ws('Candidates', CANDIDATES_COLS)
    handle_col = _col_index(ws, 'handle')
    camp_col = _col_index(ws, 'campaign')
    status_col = _col_index(ws, 'audit_status')

    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=handle_col).value == handle and \
           ws.cell(row=r, column=camp_col).value == campaign:
            ws.cell(row=r, column=status_col, value=status)
            break

    wb.save(XLSX_PATH)
    wb.close()


def get_pending_candidates(campaign: str) -> list[dict]:
    if not XLSX_PATH.exists():
        return []
    wb = load_workbook(XLSX_PATH, read_only=True)
    if 'Candidates' not in wb.sheetnames:
        wb.close()
        return []
    ws = wb['Candidates']
    camp_col = _col_index(ws, 'campaign')
    status_col = _col_index(ws, 'audit_status')
    results = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=camp_col).value == campaign and \
           ws.cell(row=r, column=status_col).value in (None, '', 'pending'):
            results.append(_row_to_dict(ws, r))
    wb.close()
    return results


# ---------------------------------------------------------------------------
# Influencers
# ---------------------------------------------------------------------------

def append_influencer(row: dict):
    """Append to Influencers sheet; skip if handle+campaign already exists."""
    wb, ws = _get_wb_ws('Influencers', INFLUENCERS_COLS)
    handle_col = _col_index(ws, 'handle') or 1
    camp_col = _col_index(ws, 'campaign') or 9
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=handle_col).value == row.get('handle') and \
           ws.cell(row=r, column=camp_col).value == row.get('campaign'):
            wb.close()
            return
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    _append_row(ws, headers, row)
    wb.save(XLSX_PATH)
    wb.close()


def update_influencer(handle: str, campaign: str, updates: dict):
    """Update specific columns for an existing influencer row."""
    wb, ws = _get_wb_ws('Influencers', INFLUENCERS_COLS)
    handle_col = _col_index(ws, 'handle')
    camp_col = _col_index(ws, 'campaign')

    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=handle_col).value == handle and \
           ws.cell(row=r, column=camp_col).value == campaign:
            for key, val in updates.items():
                col = _col_index(ws, key)
                if col:
                    ws.cell(row=r, column=col, value=val)
            break

    wb.save(XLSX_PATH)
    wb.close()


def get_influencers(campaign: str | None = None) -> list[dict]:
    """Read Influencers sheet, optionally filter by campaign."""
    if not XLSX_PATH.exists():
        return []
    wb = load_workbook(XLSX_PATH, read_only=True)
    if 'Influencers' not in wb.sheetnames:
        wb.close()
        return []
    ws = wb['Influencers']
    camp_col = _col_index(ws, 'campaign')
    results = []
    for r in range(2, ws.max_row + 1):
        if campaign and ws.cell(row=r, column=camp_col).value != campaign:
            continue
        row = _row_to_dict(ws, r)
        if any(row.values()):
            results.append(row)
    wb.close()
    return results


# ---------------------------------------------------------------------------
# Search Log
# ---------------------------------------------------------------------------

def append_search_log(row: dict):
    wb, ws = _get_wb_ws('Search Log', SEARCH_LOG_COLS)
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    _append_row(ws, headers, row)
    wb.save(XLSX_PATH)
    wb.close()


# ---------------------------------------------------------------------------
# Similar Users
# ---------------------------------------------------------------------------

def append_similar_users(rows: list[dict]) -> int:
    """Append to Similar Users sheet; dedup on (queried_handle, similar_handle)."""
    wb, ws = _get_wb_ws('Similar Users', SIMILAR_USERS_COLS)
    existing = set()
    q_col = _col_index(ws, 'queried_handle') or 1
    s_col = _col_index(ws, 'similar_handle') or 2
    for r in range(2, ws.max_row + 1):
        q = ws.cell(row=r, column=q_col).value
        s = ws.cell(row=r, column=s_col).value
        if q and s:
            existing.add((q, s))

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    added = 0
    for row in rows:
        key = (row.get('queried_handle'), row.get('similar_handle'))
        if key in existing:
            continue
        _append_row(ws, headers, row)
        existing.add(key)
        added += 1

    wb.save(XLSX_PATH)
    wb.close()
    return added


# ---------------------------------------------------------------------------
# Outreach
# ---------------------------------------------------------------------------

def append_outreach_log(row: dict):
    wb, ws = _get_wb_ws('Outreach', OUTREACH_COLS)
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    _append_row(ws, headers, row)
    wb.save(XLSX_PATH)
    wb.close()


def get_sent_handles(campaign: str) -> set[str]:
    """Return set of handles already sent outreach for this campaign."""
    if not XLSX_PATH.exists():
        return set()
    wb = load_workbook(XLSX_PATH, read_only=True)
    if 'Outreach' not in wb.sheetnames:
        wb.close()
        return set()
    ws = wb['Outreach']
    handle_col = _col_index(ws, 'handle')
    camp_col = _col_index(ws, 'campaign')
    status_col = _col_index(ws, 'status')
    sent = set()
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=camp_col).value == campaign and \
           ws.cell(row=r, column=status_col).value == 'sent':
            h = ws.cell(row=r, column=handle_col).value
            if h:
                sent.add(h)
    wb.close()
    return sent


# ---------------------------------------------------------------------------
# Keywords (context/campaigns/<name>/keywords.md)
# ---------------------------------------------------------------------------

def read_keywords(campaign_name: str) -> list[dict]:
    """Read keywords.md table, return list of {keyword, status, source, date} dicts."""
    kw_path = PROJECT_ROOT / 'context' / 'campaigns' / campaign_name / 'keywords.md'
    if not kw_path.exists():
        return []
    lines = kw_path.read_text().splitlines()
    keywords = []
    headers = []
    for line in lines:
        line = line.strip()
        if not line.startswith('|'):
            continue
        cols = [c.strip() for c in line.strip('|').split('|')]
        if not headers:
            headers = cols
            continue
        if set(line.replace('|', '').replace('-', '').replace(' ', '')) == set():
            continue  # separator row
        row = dict(zip(headers, cols))
        keywords.append(row)
    return keywords


def append_keyword(campaign_name: str, keyword: str, source: str = 'manual') -> bool:
    """Append a new pending keyword row to keywords.md.
    Returns False if keyword already exists (any status), True if appended."""
    kw_path = PROJECT_ROOT / 'context' / 'campaigns' / campaign_name / 'keywords.md'
    if not kw_path.exists():
        return False

    existing = read_keywords(campaign_name)
    for row in existing:
        if row.get('keyword', '').strip().lower() == keyword.strip().lower():
            return False

    today = date.today().isoformat()
    lines = kw_path.read_text().splitlines()

    last_table_line = None
    for i, line in enumerate(lines):
        if line.strip().startswith('|'):
            last_table_line = i

    new_row = f'| {keyword} | pending | {source} | {today} |'

    if last_table_line is None:
        kw_path.write_text('\n'.join(lines) + f'\n{new_row}\n')
    else:
        lines.insert(last_table_line + 1, new_row)
        kw_path.write_text('\n'.join(lines) + '\n')
    return True


def mark_keyword_searched(campaign_name: str, keyword: str):
    """Update keyword status to 'searched' in keywords.md."""
    kw_path = PROJECT_ROOT / 'context' / 'campaigns' / campaign_name / 'keywords.md'
    if not kw_path.exists():
        return
    lines = kw_path.read_text().splitlines()
    for i, line in enumerate(lines):
        if line.strip().startswith('|') and keyword in line and 'pending' in line:
            lines[i] = line.replace('pending', 'searched', 1)
            break
    kw_path.write_text('\n'.join(lines) + '\n')
